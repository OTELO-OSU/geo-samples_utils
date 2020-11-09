#!/usr/bin/env python3
#
#    Copyright (C) 2020 phscheffer@univ-lorraine.fr
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.
#
# -*- coding: utf-8 -*-
#
# Programme: reformatage_2_geosample.py
#
# le nom du fichier à traiter est passé en paramètre

# EN ENTREE : lecture d'un fichier excel qui contient tous les échantillons
# et ses analyses (une ligne par analyse)
# Les premières lignes contiennent des informations globales qui
#       serviront à remplir les fichiers de metadata
# Les analyses commencent à la première ligne dont la première colonne
#       contient la valeur SAMPLE_NAME, cette ligne contient le nom des champs
# La ligne suivante contient les unités des valeurs
# La colonne 1 des lignes d'analyse contient le nom des échantillons
#
# EN SORTIE : un fichier des metadatas + un fichier des données par échantillon
#
# 2020-01-20 PhS creation
# 2020-01-27 PhS - traitement de la ligne des unités
# 2020-11-09 PhS - refonte du traitement


import os, sys, time, configparser, pathlib

if sys.version_info<(3,5,3):
  sys.stderr.write("You need python 3.5.3 or later to run this script\n")
  exit(1)

from openpyxl import Workbook
from openpyxl import load_workbook
from operator import itemgetter

# ------------------------------------------------------------
# ------------------------------------------------------------

def showEnTetes():
        print("\nT_zones_obligatoires\n")
        for i in range(0,len(T_zones_obligatoires)):
                try:
                        ind = l_entete.index(T_zones_obligatoires[i])
                except (ValueError):
                        print(str(i) + "  :  " + str(T_zones_obligatoires[i]) + " : ind(N/A)")
                        continue
                else:
                        if T_zones_obligatoires[i] in T_zones_multiples:
                            if T_zones_obligatoires[i] in T_zones_multiples_z:
                                #for j,v in enumerate(T_zones_multiples_z[T_zones_obligatoires[i]]):
                                #    print(str(i) + "  :  " + str(T_zones_obligatoires[i]) + " : ind(" + str(v) + ")")
                                print(str(i) + "  :  " + str(T_zones_obligatoires[i]) + " : ind(" + str(T_zones_multiples_z[T_zones_obligatoires[i]]) + ")")
                        else:
                            print(str(i) + "  :  " + str(T_zones_obligatoires[i]) + " : ind(" + str(l_entete.index(T_zones_obligatoires[i])) + ")")
        
        print("\nT_zones_multiples\n")
        j = 0
        for i in range(len(l_entete)):
            if l_entete[i] in T_zones_multiples:
                print(str(j) + '  :  ' + T_zones_multiples[T_zones_multiples.index(l_entete[i])] + '  : ind(' + str(i) + ')')
                j += 1

        print('\nT_col_a_suppr\n')
        print(T_col_a_suppr)

        print("\nEn-tetes\n")
        for i in range(0,len(l_entete)):
                try:
                        ind = T_zones_obligatoires.index(l_entete[i])
                except (ValueError):
                        print(str(i) + "  :  " + str(l_entete[i]) + " : ind(N/A)")
                        continue
                else:
                        print(str(i) + "  :  " + str(l_entete[i]) + " : ind(" + str(T_zones_obligatoires.index(l_entete[i])) + ")")

# ---------- fin showEnTetes

# ------------------------------------------------------------

def sub_list_echt(liste, nom):
    # retourne les bornes de l'intervalle des analyses d'un
    # échantillon
    bornes= [0, 0]
    if nom == None:
        bornes[0] = bornes[1] = -1
        return bornes
    bornes[0] = bornes[1] = liste.index(nom)
    for i in range(bornes[0]+1, len(liste)):
        if liste[i] == nom:
            bornes[1] = i
        else:
            break

    return bornes

# ---------- fin de sub_list_echt

def isZoneMultiple(name):
    if name in T_zones_multiples:
        return True
    return False
# ---------- fin de IsZoneMultiple

def isZoneGroupeMultiple(name):
    # le nom est-il une sous-chaine d'une zone multiple ?
    for i in range(len(T_zones_multiples)):
        if T_zones_multiples[i] == name[0:len(T_zones_multiples[i])]:
            return T_zones_multiples[i]
    return ""
# ---------- fin de IsZoneGroupeMultiple

def msg_lig_col(enTete, lig, col, msg):
    print(str(enTete) + " : " + " ligne = " + str(lig) + ", colonne = " + str(col) + " : " + str(msg))

# ---------- fin de msg_lig_col

# ------------------------------------------------------------
# ------------------------------------------------------------

debug = 0
input_file = ""

for arg in sys.argv:
        if arg == sys.argv[0]:
                continue
        if arg == '-h' or arg == '--help':
                print("Usage : python3 " + sys.argv[0] + "-h|--help|<file> [OPTIONS]")
                print()
                print("        OPTIONS :")
                print("        --debug : print information about the execution")
                print("                  can be specified many times for get")
                print("                  more information about the execution")
                print("         -d<n>  : set debug level to n")
                print()
                print("        program reads paramaters from configuration file ./config.ini")
                print()
                print("        Example")
                print()
                print("        [GENERAL]")
                print("        ; répertoire des fichiers d'échantillons et de metadata")
                print("        rep_data = ./DATA")
                print("        ; suffixe du nom des fichiers de metadata")
                print("        suff_meta = META")
                print("        ; valeur du 1er champ des en-tetes")
                print("        sep_entete = SAMPLE_NAME")
                print("        ; mode debug (0 | 1 | 2)")
                print("        debug = 0")
                print("        [ZONES]")
                print("        zones_obligatoires=TITLE,")
                print("                           DESCRIPTION,")
                print("        zones_multiples=KEYWORD")
                print("                           INSTITUTION,")
                print()
                print("        [sampling_point_header]")
                print("        ; liste des zones du point d'échantillonnage et leur en-tête associé pour le fichier de metadata")
                print("        ; les colonnes seront dans l'ordre d'apparition des déclarations")
                print("        SAMPLING_POINT-NAME = Sampling point")
                print("        SAMPLING_POINT-COORDINATE_SYSTEM = Coordinate system")
                print("        SAMPLING_POINT-ABBREV = Abbreviation")
                print("        SAMPLING_POINT-LONGITUDE = Longitude")
                print("        SAMPLING_POINT-LATITUDE = Latitude")
                print("        SAMPLING_POINT-ELEVATION = Elevation")
                print("        SAMPLING_POINT-DESCRIPTION = Description")
                print()
                print("        [measurement_header]")
                print("        ; liste des zones de measurement et leur en-tête associé pour le fichier de metadata")
                print("        ; les colonnes seront dans l'ordre d'apparition des déclarations")
                print("        MEASUREMENT-NAME = Nature of measurement")
                print("        MEASUREMENT-ABBREV = Measurement abbreviation")
                print("        MEASUREMENT-UNIT = Units")
                print()
                print("        [methodology_header]")
                print("        ; liste des en-têtes des méthodologies")
                print("        METHODOLOGY_SAMPLING = METHODOLOGY SAMPLING")
                print("        METHODOLOGY_INSTRUMENT = METHODOLOGY INSTRUMENT")
                print("        METHODOLOGY_COMMENT = COMMENT")
                print()
                sys.exit(0)
        elif arg == '--debug':
                debug += 1
        elif arg[0:2] == '-d':
                debug += int(arg[2])
        else:
                input_file = arg

if input_file == "":
    print(sys.argv[0] + " : ERROR : file not found")
    sys.exit(1)

if not os.path.exists(input_file):
    print(sys.argv[0] + " : " + input_file + " : file not found")
    sys.exit(3)


# ------------------------------------------------------------
# répertoire d'écriture des fichiers d'échantillon
dest_rep = './DATA'
# ------------------------------------------------------------
# ------------------------------------------------------------
# --------- lecture des paramètres ---------

config = configparser.ConfigParser(allow_no_value=True)
config.optionxform = lambda option: option # conservation de la casse des clefs
config.read("config.ini")

section='GENERAL'
# répertoire des fichiers d'échantillons et de metadata
rep_data = config.get(section, 'rep_data', fallback='./DATA')
pathlib.Path(rep_data).mkdir(parents=True, exist_ok=True)

# valeur du 1er champ des en-tetes
sep_entete = config.get(section, 'sep_entete', fallback="SAMPLE_NAME")

# suffixe du nom des fichiers de metadata
suff_meta = config.get(section, 'suff_meta', fallback="META")

# nombre de lignes d'en-têtes
lig_entete = int(config.get(section, 'lig_entete', fallback=1))

# feuille active
active_sheet = config.get(section, 'active_sheet', fallback="").strip('"\'')

# mode debug
if debug == 0:
    debug = int(config.get(section, 'debug', fallback=0))

# liste des zones obligatoires dans le fichier d'entrée
temp = config.get('ZONES', "zones_obligatoires").replace("\n", "")
temp = temp.replace("'", "") # suppression de caractères d'encadrement de chaînes
temp = temp.replace('"', "") # suppression de caractères d'encadrement de chaînes
T_zones_obligatoires = temp.split(",")
# suppression des espaces devant et derrière les valeurs
for i in range(0,len(T_zones_obligatoires) - 1): T_zones_obligatoires[i] = T_zones_obligatoires[i].strip()

# valeur par défaut de zones obligatoires
temp = config.get('ZONES', "zones_obligatoires_defaut").splitlines()
T_zones_obligatoires_defaut = {}
for i in temp:
        t = i.split("=")
        T_zones_obligatoires_defaut[t[0]] = t[1]

# liste des zones multiples
temp = config.get('ZONES', "zones_multiples").replace("\n", "").strip('"\'')
T_zones_multiples = temp.split(",")
# suppression des espaces devant et derrière les valeurs
for i in range(0,len(T_zones_multiples) - 1): T_zones_multiples[i] = T_zones_multiples[i].strip()
# table des listes d'index des zones multiples
T_zones_multiples_z = {}

# liste des libellés de conversion de zones
temp = config.get('ZONES', "conv_entete_zones", fallback='').splitlines()
T_conv_zones = {}
for i in temp:
        t = i.split("=")
        T_conv_zones[t[0]] = t[1]

# liste des zones valorisées par d'autres zones
temp = config.get('ZONES', "val_zones_by_zones", fallback='').splitlines()
T_val_zones_by_zones = {}
for i in temp:
        t = i.split("=")
        T_val_zones_by_zones[t[0]] = t[1]

# les zones sont lues dans le fichier de configuration
# à chaque zone est associée un header d'affichage
T_measurement_header_z = [] # liste des zones de measurement
T_measurement_header = []   # liste des en-têtes des zones de measurement
for options in config.options('measurement_header'):
        T_measurement_header_z.append(options)
        T_measurement_header.append(config.get('measurement_header', options))

T_measurement_header_z.insert(0, "") # première colonne vide
T_measurement_header.insert(0, "") # première colonne vide

# liste des en-têtes des points d'échantillonnage
#temp = config.get('ZONES', "sampling_point_header").replace("\n", "")
#temp = temp.replace("'", "") # suppression de caractères d'encadrement de chaînes
#temp = temp.replace('"', "") # suppression de caractères d'encadrement de chaînes
#T_sampling_point_header = temp.split(",")
# suppression des espaces devant et derrière les valeurs
#for i in range(0,len(T_sampling_point_header)): T_sampling_point_header[i] = T_sampling_point_header[i].strip()
#T_sampling_point_header.insert(0, "") # première colonne vide

T_sampling_point_header_z = [] # liste des zones de sampling_point
T_sampling_point_header = []   # liste des en-têtes des zones de sampling_point
for options in config.options('sampling_point_header'):
        T_sampling_point_header_z.append(options)
        T_sampling_point_header.append(config.get('sampling_point_header', options))

T_sampling_point_header_z.insert(0, "") # première colonne vide
T_sampling_point_header.insert(0, "") # première colonne vide

T_methodology_header_z = [] # liste des zones de methodology
T_methodology_header = []   # liste des en-têtes des zones de methodology
for options in config.options('methodology_header'):
        T_methodology_header_z.append(options)
        T_methodology_header.append(config.get('methodology_header', options))

T_methodology_header_z.insert(0, "") # première colonne vide
T_methodology_header.insert(0, "") # première colonne vide

# libellé de la sous-zone de methodologie libre
methodology2_subzone_header = config.get('ZONES', 'methodology2_subzone_header', fallback='undefined')
T_methodology2_subzone = [] # liste des valeurs de methodology2

# chemins des fichiers images ou données à corriger ou insérer
# dans les noms de fichier
T_paths = {}
section = 'PATHS'
for option in config[section]:
    T_paths[option] = config.get(section, option).rstrip('/')

# suffixe obligatoire des fichiers image
T_suff_filename = {}
section = 'SUFF_FICHIER'
for option in config[section]:
    T_suff_filename[option] = config.get(section, option)


# table des index des colonnes à supprimer des fichiers de données
# zones obligatoires sauf SAMPLE_NAME
T_col_a_suppr = []

if debug > 0:
        print('T_zones_obligatoires')
        print(T_zones_obligatoires)
        print()
        print('T_zones_obligatoires_defaut')
        print(T_zones_obligatoires_defaut)
        print()
        print('T_zones_multiples')
        print(T_zones_multiples)
        print()
        print('T_conv_zones')
        print(T_conv_zones)
        print()
        print('T_val_zones_by_zones')
        print(T_val_zones_by_zones)
        print()
        print('T_sampling_point_header_z')
        print(T_sampling_point_header_z)
        print()
        print('T_sampling_point_header')
        print(T_sampling_point_header)
        print()
        print('T_measurement_header_z')
        print(T_measurement_header_z)
        print()
        print('T_measurement_header')
        print(T_measurement_header)
        print()
        print('T_methodology_header_z')
        print(T_methodology_header_z)
        print()
        print('T_methodology_header')
        print(T_methodology_header)
        print()
        print('methodology2_subzone_header = ' + methodology2_subzone_header)
        print()
        print('T_paths')
        print(T_paths)
        print()
        print('T_suff_filename')
        print(T_suff_filename)
        print()
        if debug == 1:
                sys.exit(1)

# ------------------------------------------------------------
# ------------------------------------------------------------

# fichier d'entrée
wb_input = load_workbook(filename = input_file)

# feuille principale
if active_sheet == "":
    ws_input = wb_input.active      
else:
    if active_sheet in wb_input:
        ws_input = wb_input[active_sheet]
    else:
        print("ERREUR: la feuille '" + active_sheet + "' n'existe pas dans " + str(wb_input.sheetnames))
        sys.exit(2)

col_1 = ws_input['A']

print("feuille active     : " + active_sheet)
print("nombre de lignes   : " + str(len(col_1)))
print("nombre de colonnes : " + str(ws_input.max_column))

row=1
col=1
lig_1_echt = 3 # 1ère ligne d'échantillon

# Lignes d'en-têtes (lignes 0 et 1)
l_entete = []
l_entete_unit = []
l_corres_zonObl_entete = []

# ------------------------------------------------------------
# ------------------------------------------------------------
# les deux lignes d'en-tête

for i in range(ws_input.min_column, ws_input.max_column + 1) :
        l_entete.append(ws_input.cell(row=1,column=i).value)
        if lig_entete == 2:
                l_entete_unit.append(ws_input.cell(row=2,column=i).value)

if debug == 2:
    showEnTetes()
    #sys.exit(1)

# ------------------------------------------------------------
# ------------------------------------------------------------

# ---------- MISE EN CONFORMITE AU FORMAT ATTENDU ----------

# ---------- remplacement des en-têtes non conformes
if debug >= 2:
        print("Conv zone : ")

for k, v in T_conv_zones.items():
        if k in l_entete:
                ind_conv = l_entete.index(k)
                if debug >= 2:
                        print("Conv zone : " + str(k))
                        print("Conv zone : " + l_entete[ind_conv] + " -> " + v)
                # modification de l'en-tête dans la liste
                l_entete[ind_conv] = v
                # modification de l'en-tête dans la worksheet
                ws_input.cell(row=1, column=ind_conv+1, value=v)

# ---------- SAMPLE_NAME
if sep_entete in l_entete:
        sni = l_entete.index(sep_entete) + 1
        if sni != 1:
                # deplacement de la colonne SAMPLE_NAME en position 1
                if debug >= 2:
                        showEnTetes()
                        print("Déplacement de la colonne " + str(sni) + " en colonne 1")
                # -- insertion d'une première colonne vide
                ws_input.insert_cols(1)
                sni += 1
                # -- copie des valeurs des cellules
                for i in range(1, ws_input.max_row+1):
                        #print("case[" + str(i) + "," + str(sni) + "] = " + str(ws_input.cell(row=i, column=sni).value))
                        ws_input.cell(row=i, column=1, value=ws_input.cell(row=i, column=sni).value)
                # -- suppression de la colonne d'origine
                ws_input.delete_cols(sni)
                # -- relecture des en-têtes
                l_entete = []
                for i in range(ws_input.min_column, ws_input.max_column + 1) :
                        l_entete.append(ws_input.cell(row=1,column=i).value)
                if debug >= 2:
                        showEnTetes()
else:
        print("index " + sep_entete + " MANQUANT")
        sys.exit(2)

# ---------- ZONES MULTIPLES

for i in range(ws_input.min_column, ws_input.max_column + 1) :
        if ws_input.cell(row=1,column=i).value in T_zones_multiples:
            # zone multiple
            if ws_input.cell(row=1,column=i).value in T_zones_multiples_z:
                # ajout de colonne pour la zone existante
                T_zones_multiples_z[ws_input.cell(row=1,column=i).value].append(i)
            else:
                # création de la zone dans la table
                T_zones_multiples_z[ws_input.cell(row=1,column=i).value] = []
                T_zones_multiples_z[ws_input.cell(row=1,column=i).value].append(i)
        else:
            # zone de groupe multiple (ex: REFERENT -> REFERENT_NAME, REFERENT_FIRST_NAME, REFERENT_MAIL)
            for j in range(len(T_zones_multiples)):
                if ws_input.cell(row=1,column=i).value[0:len(T_zones_multiples[j])] == T_zones_multiples[j]:
                    # le début de la zone correspond  à une zone multiple
                    if T_zones_multiples[j] in T_zones_multiples_z:
                        # ajout de colonne pour la zone existante
                        T_zones_multiples_z[T_zones_multiples[j]].append(i)
                    else:
                        # création de la zone dans la table
                        T_zones_multiples_z[T_zones_multiples[j]] = []
                        T_zones_multiples_z[T_zones_multiples[j]].append(i)
                    break

if debug >= 4:
        print()
        print('T_zones_multiples')
        print(T_zones_multiples)
        print()
        print('T_zones_multiples_z')
        print(T_zones_multiples_z)
        print()


# ---------- TABLE DES COLONNES A SUPPRIMER DANS LES FICHIERS DE DONNEES

# zones obligatoires sauf SAMPLE_NAME
T_col_a_suppr_temp = []
for i in range(0,len(T_zones_obligatoires)):
    if T_zones_obligatoires[i] == "ANALYST":
        print("COLONNES A SUPPRIMER : " + T_zones_obligatoires[i])
    if T_zones_obligatoires[i] != sep_entete:
        try:
            ind = l_entete.index(T_zones_obligatoires[i])
        except (ValueError):
            if isZoneGroupeMultiple(T_zones_obligatoires[i]) in T_zones_multiples_z:
                T_col_a_suppr_temp += T_zones_multiples_z[isZoneGroupeMultiple(T_zones_obligatoires[i])]
            continue
        else:
            if T_zones_obligatoires[i] in T_zones_multiples:
                if T_zones_obligatoires[i] in T_zones_multiples_z:
                    T_col_a_suppr_temp += T_zones_multiples_z[T_zones_obligatoires[i]]
            else:
                T_col_a_suppr_temp.append(ind+1)

T_col_a_suppr_temp.sort(reverse=True)

# déduplication de T_col_a_suppr
for i in T_col_a_suppr_temp:
    if i not in T_col_a_suppr:
        T_col_a_suppr.append(i)

if debug >= 4:
        print()
        print('T_col_a_suppr_temp')
        print(T_col_a_suppr_temp)

# ------------------------------------------------------------

if debug >= 2:
    showEnTetes()
if debug == 2:
    sys.exit(1)

# ------------------------------------------------------------
# ------------------------------------------------------------

# liste de lecture triée
if 'MEASUREMENT-ABBREV' not in l_entete:
    print("ERREUR: colonne " + 'MEASUREMENT-ABBREV' + " manquante")
    sys.exit(2)

liste_tri = []
liste_tri_elem = ["", "", ""]
colMeasurementAbbrev = l_entete.index('MEASUREMENT-ABBREV') + 1
if debug >= 3:
    print('colMeasurementAbbrev = ' + str(colMeasurementAbbrev))

#for i in range(input.min_column, ws_input.max_column + 1) :
for i in range(lig_entete+1,len(col_1)+1):
    liste_tri_elem[0] = ws_input.cell(row=i, column=1).value # SAMPLE_NAME
    liste_tri_elem[1] = ws_input.cell(row=i, column=colMeasurementAbbrev).value # MEASUREMENT-ABBREV
    liste_tri_elem[2] = i # numero ligne
    #print(liste_tri_elem)
    liste_tri.append([''] * 3)
    liste_tri[len(liste_tri) - 1][0] = str(liste_tri_elem[0])
    liste_tri[len(liste_tri) - 1][1] = str(liste_tri_elem[1])
    liste_tri[len(liste_tri) - 1][2] = liste_tri_elem[2]

liste_tri.sort(key=itemgetter(0,1), reverse=False)

liste_echt_tri = []
for i in liste_tri:
    liste_echt_tri.append(i[0])

if debug > 0:
    print('liste triee')
    for i,v in enumerate(liste_tri):
        if debug > 0 and debug <=3 and i >= 18:
            break
        print(str(i) + ' : ' + str(liste_tri[i]))

# ------------------------------------------------------------

# ---------- MAIN ----------

# ---------- traitement des échantillons ----------

# initialisation pour le premier échantillon
atime = time.perf_counter()

lig = liste_tri[0][2]
nom_echt = liste_tri[0][0]
bornes_echt_sav = []
bornes_echt = []
while True:
        
        # recherche de l'intervalle des enregistrements de l'échantillon
        n_echt = ws_input.cell(row=lig, column=ws_input.min_column).value
        bornes_echt_sav = bornes_echt
        bornes_echt = sub_list_echt(liste_echt_tri, n_echt)
        if bornes_echt[0] == -1:
            # ERREUR sur le nom de l'échantillon
            print('ERREUR: lig ' + str(lig) + 'nom echt : ' + str(n_echt))
            # positionnement sur le nom suivant
            lig = liste_tri[bornes_echt_sav[1] + 2][2]
            nom_echt = liste_tri[bornes_echt_sav[1] + 2][0]
            continue

        # =====================================================
        #            les zones de METADATA
        # =====================================================

        l = 1 # indice de ligne dans le nouveau fichier

        # initialisation du nouveau fichier de metadata
        wb_metadata = Workbook()
        ws_metadata = wb_metadata.active

        # les zones de METADATA
        for i in range(0,len(T_zones_obligatoires)):
                if (T_zones_obligatoires[i] not in T_measurement_header_z
                    and T_zones_obligatoires[i] not in T_sampling_point_header_z
                    and T_zones_obligatoires[i] not in T_methodology_header_z
                    and T_zones_obligatoires[i] != 'METHODOLOGY2'):

                    # traitement des zones multiples
                    zoneObligMultiple = ""
                    if isZoneMultiple(T_zones_obligatoires[i]) and T_zones_obligatoires[i] in T_zones_multiples_z:
                        zoneObligMultiple = T_zones_obligatoires[i]
                    elif isZoneGroupeMultiple(T_zones_obligatoires[i]) in T_zones_multiples_z:
                        zoneObligMultiple = isZoneGroupeMultiple(T_zones_obligatoires[i])
                    if debug >= 4:
                        print("GROUPE ZONE MULTIPLE : " + zoneObligMultiple + "(" + T_zones_obligatoires[i] + ")");
                    if zoneObligMultiple != "":
                        if debug >= 3:
                            print('ZONE MULTIPLE: ' + T_zones_obligatoires[i] + ' : colonnes = ' + str(T_zones_multiples_z[zoneObligMultiple]))
                        for j in T_zones_multiples_z[zoneObligMultiple]:
                            if debug >= 3:
                                print('ZONE MULTIPLE: ' + zoneObligMultiple + ' : colonne = ' + str(j))
                            if ws_input.cell(row=lig, column=j).value != None:
                                #ws_metadata.cell(row=l, column=1, value=T_zones_obligatoires[i])
                                ws_metadata.cell(row=l, column=1, value=ws_input.cell(row=1, column=j).value)
                                if T_zones_obligatoires[i] in T_paths:
                                    # substitution de chemin
                                    ind_path = int(ws_input.cell(row=lig, column=j).value.rfind('/'))
                                    ws_metadata.cell(row=l, column=2, value=T_paths[T_zones_obligatoires[i]] + '/' + ws_input.cell(row=lig, column=j).value[ind_path+1:])
                                else:
                                    ws_metadata.cell(row=l, column=2, value=ws_input.cell(row=lig, column=j).value)
                                if T_zones_obligatoires[i] in T_suff_filename:
                                    # insertion du suffixe si manquant
                                    if T_suff_filename[T_zones_obligatoires[i]] not in ws_metadata.cell(row=l, column=2).value:
                                        try:
                                            ind_suff = str(ws_metadata.cell(row=l, column=2).value).index('.')
                                        except(ValueError):
                                            if debug >= 3:
                                                print("SUFFIXE : caractère '.' non trouvé dans le nom du fichier " + str(ws_metadata.cell(row=l, column=2)))
                                        else:
                                            val_suff_temp = ws_metadata.cell(row=l, column=2).value
                                            if debug > 4:
                                                print("SUFFIXE: METADATA : " + val_suff_temp + ' index ext = ' + str(ind_suff))
                                            ws_metadata.cell(row=l, column=2, value=val_suff_temp[0:ind_suff] + T_suff_filename[T_zones_obligatoires[i]] + val_suff_temp[ind_suff:])
                                    if debug > 4:
                                        print("SUFFIXE: METADATA : " + ws_metadata.cell(row=1, column=1).value + ' = ' + ws_metadata.cell(row=l, column=2).value)
                                l += 1

                    else:
                        ws_metadata.cell(row=l, column=1, value=T_zones_obligatoires[i])

                        # valorisation de zones par d'autres zones
                        if T_zones_obligatoires[i] in T_val_zones_by_zones:
                            if debug >= 3:
                                    print(str(lig) + " : ZONE " +  T_zones_obligatoires[i])
                                    print(str(lig) + " : ZONE target " + T_val_zones_by_zones[T_zones_obligatoires[i]])
                                    print(str(lig) + " : indice de ZONE target " + str(l_entete.index(T_val_zones_by_zones[T_zones_obligatoires[i]])))
                                    print(str(lig) + " : valeur de ZONE [A" + str(lig) + "] " + str(ws_input.cell(row=lig, column=1).value))
                                    print(str(lig) + " : valeur de ZONE target " + str(ws_input.cell(row=lig, column=l_entete.index(T_val_zones_by_zones[T_zones_obligatoires[i]])+1).value))
                            ws_metadata.cell(row=l, column=2, value=ws_input.cell(row=lig, column=l_entete.index(T_val_zones_by_zones[T_zones_obligatoires[i]])+1).value)
                        else:

                            try:
                                ind = l_entete.index(T_zones_obligatoires[i])
                            except (ValueError):
                                # colonne manquante
                                if T_zones_obligatoires[i] in T_zones_obligatoires_defaut:
                                        ws_metadata.cell(row=l, column=2, value=T_zones_obligatoires_defaut[T_zones_obligatoires[i]])
                                else:
                                        ws_metadata.cell(row=l, column=2, value='N/A')
                                l += 1
                                continue
                            else:
                                ws_metadata.cell(row=l, column=2, value=ws_input.cell(row=lig,column=ind+1).value)
                                if ws_metadata.cell(row=l, column=2).value == None:
                                    # zone vide
                                    if T_zones_obligatoires[i] in T_zones_obligatoires_defaut:
                                        if debug >= 3:
                                            print('ZONE: ' + T_zones_obligatoires[i] + ' : valeur par défaut : (' + T_zones_obligatoires_defaut[T_zones_obligatoires[i]] + ')')
                                        ws_metadata.cell(row=l, column=2, value=T_zones_obligatoires_defaut[T_zones_obligatoires[i]])
                                    else:
                                        ws_metadata.cell(row=l, column=2, value="-----")

                                    if '_DATE' in ws_metadata.cell(row=l, column=1).value:
                                            # valeur par défaut pour zone de date vide
                                            ws_metadata.cell(row=l, column=2, value="3000-01-01")
                                else:
                                    if T_zones_obligatoires[i] in T_paths:
                                        # substitution de chemin
                                        ind_path = int(ws_input.cell(row=l, column=2).value.rfind('/'))
                                        ws_metadata.cell(row=l, column=2, value=T_paths[T_zones_obligatoires[i]] + '/' + ws_input.cell(row=l, column=2).value[ind_path+1:])

                                    if '_DATE' in ws_metadata.cell(row=l, column=1).value:
                                        # suppression d'eventuelles informations d'heure
                                        if debug >= 3:
                                            print('DATE : (lig=' + str(l) + ') ' + ws_metadata.cell(row=l, column=1).value + ' = ' + str(ws_metadata.cell(row=l, column=2).value))
                                        temp_date = str(ws_metadata.cell(row=l, column=2).value)
                                        try:
                                            i_date = temp_date.index(" ")
                                        except(ValueError):
                                            l += 1
                                            continue
                                        else:
                                            ws_metadata.cell(row=l, column=2, value=temp_date[0:i_date])
                                            if debug >= 3:
                                                print('DATE : ' + ws_metadata.cell(row=l, column=1).value + ' = ' + str(ws_metadata.cell(row=l, column=2).value))

                        l += 1

        # les zones de SAMPLE_POINT
        l += 1
        for i,v in enumerate(T_sampling_point_header):
            ws_metadata.cell(row=l, column=i+1, value=v)
        l += 1
        ws_metadata.cell(row=l, column=1, value='SAMPLING_POINT')
        for i in range(1,len(T_sampling_point_header_z)):
            if T_sampling_point_header_z[i] in l_entete:
                sp_col = l_entete.index(T_sampling_point_header_z[i]) + 1
                if ws_input.cell(row=lig, column=sp_col).value == None:
                    if debug >=3: print(str(T_sampling_point_header_z[i]) + ' : ' + '-----')
                    sp_value = '-----'
                else:
                    if debug >=3: print(str(T_sampling_point_header_z[i]) + ' : ' + str(ws_input.cell(row=lig, column=sp_col).value))
                    sp_value = str(ws_input.cell(row=lig, column=sp_col).value)
            else:
                if debug >=3: print(str(T_sampling_point_header_z[i]) + ' : ' + 'N/A')
                sp_value = 'N/A'
            ws_metadata.cell(row=l, column=i+1, value=sp_value)

        # =====================================================
        # parcours de la liste
        if debug > 3: print("parcours de l'échantillon")
        # initialisation de la feuille de données
        wb_output = Workbook()
        ws_output = wb_output.active
        l_don = 1
        i_methodology = 0 # indice de ligne de methodology dans la table
        T_meta_methodology = [] # table des methodologies
        i_measurement = 0 # indice de ligne de measurement dans la table
        T_meta_measurement = [] # table des measurements
        # les deux lignes d'en-tête
        for i in range(ws_input.min_column, ws_input.max_column + 1) :
            ws_output.cell(row=l_don, column=i, value=l_entete[i-1])
            if lig_entete == 2:
                ws_output.cell(row=l_don+1, column=i, value=l_entete_unit[i-1])
        
        # la ligne d'analyse de l'échantillon
        # la ligne des unités reste vide, elle est prise en compte dans le reformatage
        l_don += 2 # première ligne d'analyse dans le nouveau fichier
        analyse = liste_tri[bornes_echt[0]][1]
        if debug > 3: print("parcours de l'échantillon lignes " + str(bornes_echt[0]) + ' - ' + str(bornes_echt[1]))
        for i in range(bornes_echt[0], bornes_echt[1] + 1):
            print('ligne : ' + str(i))
            if debug >= 3:
                print('ind : ' + str(i) + '  echantillon : ' + nom_echt + '  analyse : ' + analyse)
                print(liste_tri[i])
            if analyse != liste_tri[i][1]:
                if debug > 2: print("changement d'analyse : nouvelle analyse %s analyse courante %s" % (analyse,liste_tri[i][1]))
                # changement de type d'analyse, changement de fichier
                # écriture du fichier de données

                if debug >= 3:
                    print('FIC DONNÉES : suppression des colonnes ' + str(T_col_a_suppr))
                for i_colas in T_col_a_suppr:
                    if debug >= 3:
                        print('FIC DONNÉES : suppression de la colonne ' + str(i_colas))
                    ws_output.delete_cols(i_colas)
                # suppression des colonnes sans valeur
                for i_colas in range(ws_output.max_column, ws_output.min_column - 1, -1):
                    if debug >= 5:
                        print('FIC DONNÉES : parcours des colonnes de DATA : ' + str(i_colas) + ' (' + str(ws_output.max_row) + ' lignes)')
                    b_colas = True
                    for j_rawas in range(3, ws_output.max_row + 1):
                        if debug >= 5:
                            print('FIC DONNÉES : parcours des lignes de la colonne ' + str(ws_output.cell(row=1, column=i_colas).value) + ' : (col ' + str(i_colas) + ') de DATA : lig ' + str(j_rawas))
                            if ws_output.cell(row=1, column=i_colas).value == 'CORE_AZIMUT':
                                print('FIC DONNÉES : parcours des lignes de la colonne CORE_AZIMUT (col ' + str(i_colas) + ') de DATA : ' + str(j_rawas) + ', val = ' + str(ws_output.cell(row=j_rawas, column=i_colas).value))
                        if ws_output.cell(row=j_rawas, column=i_colas).value != None:
                            b_colas = False
                    if b_colas == True:
                        if debug >= 5:
                            print('FIC DONNÉES : suppression de la colonne sans valeur : ' + str(i_colas))
                        ws_output.delete_cols(i_colas)

                dest_filename_don = dest_rep + '/' + nom_echt + '_' + analyse + '.xlsx'
                wb_output.save(filename=dest_filename_don)
                curr_time = time.perf_counter()
                if debug >= 3: print("%6.3f : Ecriture du fichier d'analyse %s" % (curr_time - atime, dest_filename_don))
                atime = curr_time
                # initialisation de la feuille de données
                wb_output = Workbook()
                ws_output = wb_output.active
                l_don = 1 # première ligne dans le nouveau fichier
                # les deux lignes d'en-tête
                for j in range(ws_input.min_column, ws_input.max_column + 1):
                    ws_output.cell(row=l_don, column=j, value=l_entete[j-1])
                    if debug > 4:
                        msg_lig_col("DATA : en-têtes", l_don, j, ws_output.cell(row=l_don, column=j).value)
                    if lig_entete == 2:
                        ws_output.cell(row=l_don+1, column=j, value=l_entete_unit[j-1])
                        if debug > 4:
                            msg_lig_col("DATA : unités", l_don, j, ws_output.cell(row=l_don, column=j).value)
                l_don += 2 # première ligne d'analyse dans le nouveau fichier
                analyse = liste_tri[i][1]
                if debug >= 3: print('analyse : ' + analyse + ' indice : ' + str(i) + '  borne de fin : ' + str(bornes_echt[1]))
            #else:
            ind_lig = liste_tri[i][2]

            # fichier de données
            #print('ligne de données : ' + str(liste_tri[i][2]) + '  measurement_id : ' + ws_input.cell(row=ind_lig,column=11).value)
            for j in range(ws_input.min_column, ws_input.max_column + 1) :
                if debug >= 3:
                    print("DATA : " + "ZONE : ligne = " + str(ind_lig) + ", colonne = " + str(j))
                    print("DATA : " + "ZONE : " + str(ws_input.cell(row=1, column=j).value) + ' = ' + str(ws_input.cell(row=ind_lig, column=j).value))

                #ws_output.cell(row=l_don, column=i, value=ws_input.cell(row=liste_tri[i][2],column=i).value)
                ws_output.cell(row=l_don, column=j, value=ws_input.cell(row=ind_lig,column=j).value)

                if (ws_output.cell(row=1, column=j).value in T_paths
                    and ws_output.cell(row=l_don, column=j).value != None):
                    # substitution de chemin
                    if debug > 4:
                        print('(' + str(ind_lig) + ') DATA : PATH : ' + ws_input.cell(row=1, column=j).value + ' VALUE = ' + str(ws_output.cell(row=l_don, column=j).value))
                        print('(' + str(ind_lig) + ') DATA : PATH : ' + T_paths[ws_output.cell(row=1, column=j).value] + ' VALUE = ' + str(ws_output.cell(row=l_don, column=j).value))

                    ind_path = int(ws_output.cell(row=l_don, column=j).value.rfind('/'))
                    ws_output.cell(row=l_don, column=j, value=T_paths[ws_output.cell(row=1, column=j).value] + '/' + ws_output.cell(row=l_don, column=j).value[ind_path+1:])
                    if debug >= 3:
                        print("DATA : " + "ZONE : " + ws_output.cell(row=1, column=j).value + ' = ' + ws_output.cell(row=l_don, column=j).value)

                if (ws_output.cell(row=1, column=j).value in T_suff_filename
                    and ws_output.cell(row=l_don, column=j).value != None):
                    # insertion du suffixe si manquant
                    if T_suff_filename[ws_output.cell(row=1, column=j).value] not in ws_output.cell(row=l_don, column=j).value:
                        try:
                            ind_suff = str(ws_output.cell(row=l_don, column=j).value).index('.')
                        except(ValueError):
                            if debug >=3:
                                print("SUFFIXE: caractère '.' non trouvé dans le nom du fichier " + ws_output.cell(row=l_don, column=j).value)
                        else:
                            ws_output.cell(row=l_don, column=j, value=ws_output.cell(row=l_don, column=j).value[0:ind_suff] + T_suff_filename[ws_output.cell(row=1, column=j).value] + ws_output.cell(row=l_don, column=j).value[ind_suff:])
                    if debug > 4:
                        print("SUFFIXE: ZONE : " + ws_output.cell(row=1, column=j).value + ' = ' + ws_output.cell(row=l_don, column=j).value)

                # suppression d'éventuelles informations d'heure
                if '_DATE' in ws_output.cell(row=1, column=j).value:
                    if debug >= 3:
                        print('DONNEES : date détectée : ' + l_entete[j-1] + ' = ' + str(ws_output.cell(row=l_don, column=j).value))
                    temp_date = str(ws_output.cell(row=l_don, column=j).value)
                    try:
                        i_date = temp_date.index(" ")
                    except(ValueError):
                        continue
                    else:
                        if debug >= 3:
                            print('DONNEES : date format non conforme : ' + l_entete[j-1] + ' = ' + str(ws_output.cell(row=l_don, column=j).value))
                        ws_output.cell(row=l_don, column=j, value=temp_date[0:i_date])
                        if debug >= 3:
                            print('DONNEES : date corrigée : ' + l_entete[j-1] + ' = ' + str(ws_output.cell(row=l_don, column=j).value))

            l_don += 1 # ligne suivante

            # ---------- zones de MEASUREMENT pour metadata
            m = []
            m.append('MEASUREMENT')
            for j in range(1,len(T_measurement_header_z)):
                if T_measurement_header_z[j] in T_val_zones_by_zones:
                    # valorisation de zones par d'autres zones
                    if debug >= 3:
                        print('MEASUREMENT')
                        print('valorisation de ' + T_measurement_header_z[j] + ' par ' + str(T_val_zones_by_zones[T_measurement_header_z[j]]))
                        print(str(ind_lig) + " : indice de ZONE target " + str(l_entete.index(T_val_zones_by_zones[T_measurement_header_z[j]])))
                        print(str(ind_lig) + " : valeur de ZONE target " + str(ws_input.cell(row=ind_lig, column=l_entete.index(T_val_zones_by_zones[T_measurement_header_z[j]])+1).value))
                    try:
                        v_m = ws_input.cell(row=ind_lig, column=l_entete.index(T_val_zones_by_zones[T_measurement_header_z[j]])+1).value
                        if v_m == None:
                            m.append('-----')
                        else:
                            m.append(v_m)
                    except(ValueError): m.append('N/A')
                else:
                    # valorisation de la zone
                    try:
                        v_m = ws_input.cell(row=ind_lig,column=l_entete.index(T_measurement_header_z[j])+1).value
                        if v_m == None:
                            m.append('-----')
                        else:
                            m.append(v_m)
                    except(ValueError): m.append('N/A')
            if m not in T_meta_measurement:
                if debug > 3: print(m)
                T_meta_measurement.append(m)
                i_measurement += 1

            # ---------- zones de METHODOLOGY pour metadata
            m = []
            m.append('METHODOLOGY')
            for j in range(1,len(T_methodology_header_z)):
                try:
                    v_m = ws_input.cell(row=ind_lig,column=l_entete.index(T_methodology_header_z[j])+1).value
                    if v_m == None:
                        m.append('-----')
                    else:
                        m.append(v_m)
                except(ValueError): m.append('N/A')
            if m not in T_meta_methodology:
                if debug > 3: print(m)
                T_meta_methodology.append(m)
                i_methodology += 1
            # ----------
            # ---------- zones de METHODOLOGY2 pour metadata
            try:
                v_m2 = l_entete.index('METHODOLOGY2')
            except(ValueError):
                continue
            else:
                if ws_input.cell(row=ind_lig,column=v_m2+1).value not in T_methodology2_subzone:
                    T_methodology2_subzone.append(ws_input.cell(row=ind_lig,column=v_m2+1).value)

        # ---------- fin du parcours des analyses

        # écriture du fichier de données
        dest_filename_don = dest_rep + '/' + nom_echt + '_' + analyse + '.xlsx'
        if debug >= 3:
            print('FIC DONNÉES : suppression des colonnes ' + str(T_col_a_suppr))
        for i_colas in T_col_a_suppr:
            if debug >= 3:
                print('FIC DONNÉES : suppression de la colonne ' + str(i_colas))
            ws_output.delete_cols(i_colas)
        # suppression des colonnes sans valeur
        for i_colas in range(ws_output.max_column, ws_output.min_column - 1, -1):
            b_colas = True
            for j_rawas in range(3, ws_output.max_row + 1):
                if ws_output.cell(row=j_rawas, column=i_colas).value != None:
                    b_colas = False
            if b_colas == True:
                if debug >= 5:
                    print('FIC DONNÉES : suppression de la colonne sans valeur : ' + str(i_colas))
                ws_output.delete_cols(i_colas)

        wb_output.save(filename=dest_filename_don)
        if debug > 0:
            curr_time = time.perf_counter()
            print("%6.3f : Fin echantillon : Ecriture du fichier d'analyse %s" % (curr_time - atime, dest_filename_don))
            atime = curr_time
        # =====================================================
        # ---------- ajout des measurements aux metadatas
        l += 2
        for i,v in enumerate(T_measurement_header):
            ws_metadata.cell(row=l, column=i+1, value=v)
        l += 1
        for v in T_meta_measurement:
            for j,w in enumerate(v):
                #print(str(v) + ' : ' + str(w))
                ws_metadata.cell(row=l, column=j+1, value=w)
            l += 1
        # ---------- ajout des methodologies aux metadatas
        l += 2
        for i,v in enumerate(T_methodology_header):
            ws_metadata.cell(row=l, column=i+1, value=v)
        l += 1
        for v in T_meta_methodology:
            for j,w in enumerate(v):
                #print(str(v) + ' : ' + str(w))
                ws_metadata.cell(row=l, column=j+1, value=w)
            l += 1
        # ---------- ajout des methodologies libres aux metadatas
        l += 1
        for i,v in enumerate(T_methodology2_subzone):
            ws_metadata.cell(row=l, column=1, value='METHODOLOGY')
            ws_metadata.cell(row=l, column=2, value=methodology2_subzone_header)
            ws_metadata.cell(row=l, column=3, value=v)
            if debug >= 3:
                print(ws_metadata.cell(row=l, column=1).value + ' , ' + ws_metadata.cell(row=l, column=2).value + ' , ' + ws_metadata.cell(row=l, column=3).value)
            l += 1
        l += 1
        # ---------- écriture du fichier de metadata
        if nom_echt != "" :
            dest_filename = dest_rep + '/' + nom_echt + '_' + suff_meta + '.xlsx'
            wb_metadata.save(filename=dest_filename)
            if debug > 0:
                curr_time = time.perf_counter()
                print('%6.3f : Ecriture du fichier %s' % (curr_time - atime, dest_filename))
                atime = curr_time

        # =====================================================
        # ---------- ARRET ?
        #print('fin de liste ? ' + str(bornes_echt[1]) + ' == ' + str(len(liste_tri) - 1))
        if bornes_echt[1] == len(liste_tri) - 1:
            sys.exit(0)
        
        # =====================================================
        # =====================================================
        # ---------- nouvel échantillon
        lig = liste_tri[bornes_echt[1] + 1][2]
        nom_echt = liste_tri[bornes_echt[1] + 1][0]
        # =====================================================


# ---------- fin du fichier ----------
